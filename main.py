import asyncio
import os
import tempfile
from datetime import datetime, date, timedelta

import aiosqlite
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    FSInputFile,
)
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext


# Railway Variables (set these in Railway -> Variables)
BOT_TOKEN = os.getenv("BOT_TOKEN")
RESET_PASSWORD = os.getenv("RESET_PASSWORD", "")  # set in Railway -> Variables

# DB file path
DB_PATH = "participants.sqlite"
# If you enabled Railway Volume mounted to /data, use this instead:
# DB_PATH = "/data/participants.sqlite"

# Put your Telegram user_ids here (2 admins supported)
ADMIN_IDS = {922603146, 123456789}  # <-- replace with real IDs


# ---------- FSM ----------
class Reg(StatesGroup):
    waiting_consent = State()
    waiting_phone = State()
    waiting_first_name = State()
    waiting_last_name = State()


class AdminFSM(StatesGroup):
    reset_wait_password = State()
    reset_confirm = State()
    export_wait_from = State()
    export_wait_to = State()
    list_wait_from = State()
    list_wait_to = State()


# ---------- DB ----------
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS participants (
    id INTEGER PRIMARY KEY AUTOINCREMENT,       -- participant number (1,2,3,...)
    telegram_id INTEGER UNIQUE NOT NULL,
    phone TEXT UNIQUE NOT NULL,                 -- unique by phone
    first_name TEXT NOT NULL,
    last_name TEXT NOT NULL,
    consent INTEGER NOT NULL,                   -- 1 = agreed
    created_at TEXT NOT NULL                    -- UTC ISO string
);
"""


async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(CREATE_TABLE_SQL)
        await db.commit()


async def get_by_telegram_id(telegram_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT id, telegram_id, phone, first_name, last_name, consent, created_at "
            "FROM participants WHERE telegram_id = ?",
            (telegram_id,),
        )
        return await cur.fetchone()


async def get_by_phone(phone: str):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT id, telegram_id, phone, first_name, last_name, consent, created_at "
            "FROM participants WHERE phone = ?",
            (phone,),
        )
        return await cur.fetchone()


async def insert_participant(telegram_id: int, phone: str, first_name: str, last_name: str, consent: int) -> int:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            """
            INSERT INTO participants (telegram_id, phone, first_name, last_name, consent, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (telegram_id, phone, first_name, last_name, consent, datetime.utcnow().isoformat()),
        )
        await db.commit()
        return cur.lastrowid


def _range_where_clause(from_iso: str | None, to_iso: str | None):
    # ISO timestamps can be compared lexicographically
    if from_iso and to_iso:
        return "WHERE created_at >= ? AND created_at < ?", (from_iso, to_iso)
    if from_iso:
        return "WHERE created_at >= ?", (from_iso,)
    if to_iso:
        return "WHERE created_at < ?", (to_iso,)
    return "", ()


async def fetch_participants(from_iso: str | None = None, to_iso: str | None = None):
    where_sql, params = _range_where_clause(from_iso, to_iso)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            f"SELECT id, telegram_id, phone, first_name, last_name, consent, created_at "
            f"FROM participants {where_sql} ORDER BY id ASC",
            params
        )
        return await cur.fetchall()


async def count_participants(from_iso: str | None = None, to_iso: str | None = None) -> int:
    where_sql, params = _range_where_clause(from_iso, to_iso)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(f"SELECT COUNT(*) FROM participants {where_sql}", params)
        (cnt,) = await cur.fetchone()
        return cnt


async def reset_database():
    # Full wipe + reset AUTOINCREMENT counter back to 1
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM participants")
        await db.execute("DELETE FROM sqlite_sequence WHERE name='participants'")
        await db.commit()


# ---------- Keyboards ----------
def user_start_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="🚀 Старт")]],
        resize_keyboard=True
    )


def consent_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="✅ Согласен"), KeyboardButton(text="❌ Не согласен")]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )


def contact_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Отправить номер телефона", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def admin_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Список"), KeyboardButton(text="📤 Экспорт")],
            [KeyboardButton(text="📤 Экспорт сегодня")],
            [KeyboardButton(text="🧹 Ресет базы")],
            [KeyboardButton(text="⬅️ Закрыть меню")]
        ],
        resize_keyboard=True
    )


def admin_reset_confirm_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="✅ Да, стереть всё"), KeyboardButton(text="❌ Отмена")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def admin_filter_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сегодня"), KeyboardButton(text="Все")],
            [KeyboardButton(text="Диапазон дат")],
            [KeyboardButton(text="⬅️ Назад")]
        ],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def admin_back_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="⬅️ Назад")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )


# ---------- Helpers ----------
def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def normalize_phone(phone: str) -> str:
    return phone.strip().replace(" ", "")


def parse_ymd(s: str) -> date | None:
    try:
        return datetime.strptime(s.strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def day_range_utc(d: date) -> tuple[str, str]:
    start = datetime(d.year, d.month, d.day)
    end = start + timedelta(days=1)
    return start.isoformat(), end.isoformat()


def range_from_args(args_text: str) -> tuple[str | None, str | None, str | None]:
    """
    Accepts:
      - "" -> no filter
      - "today" -> today's UTC day
      - "YYYY-MM-DD YYYY-MM-DD" -> inclusive date range
    Returns: (from_iso, to_iso, error)
    """
    t = (args_text or "").strip()
    if not t:
        return None, None, None

    if t.lower() == "today":
        f, to = day_range_utc(datetime.utcnow().date())
        return f, to, None

    parts = t.split()
    if len(parts) == 2:
        d1 = parse_ymd(parts[0])
        d2 = parse_ymd(parts[1])
        if not d1 or not d2:
            return None, None, "Неверный формат даты. Используй YYYY-MM-DD YYYY-MM-DD"
        f = datetime(d1.year, d1.month, d1.day).isoformat()
        to_excl_date = d2 + timedelta(days=1)
        to = datetime(to_excl_date.year, to_excl_date.month, to_excl_date.day).isoformat()
        return f, to, None

    return None, None, "Неверные аргументы. Примеры: /export today или /export 2026-02-01 2026-02-06"


def autosize_worksheet_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


async def export_to_excel_and_send(message: Message, rows, suffix: str):
    """
    Save to /tmp because Railway filesystem may be read-only in app dir.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    ws.append(["Номер", "Telegram ID", "Телефон", "Имя", "Фамилия", "Согласие", "Дата регистрации (UTC)"])
    for r in rows:
        pid, tid, phone, fn, ln, consent, created_at = r
        ws.append([pid, tid, phone, fn, ln, "Да" if consent else "Нет", created_at])

    autosize_worksheet_columns(ws)

    with tempfile.NamedTemporaryFile(prefix=f"participants_{suffix}_", suffix=".xlsx", delete=False, dir="/tmp") as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)

    try:
        await message.answer_document(
            FSInputFile(tmp_path),
            caption=f"Выгрузка участников: {len(rows)} записей"
        )
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


# ---------- List / Export core (NO fake Message!) ----------
async def send_list(message: Message, args: str):
    from_iso, to_iso, err = range_from_args(args)
    if err:
        await message.answer(err)
        return

    cnt = await count_participants(from_iso, to_iso)
    rows = await fetch_participants(from_iso, to_iso)

    preview = rows[:30]
    lines = [f"{pid}. {fn} {ln} — {phone}" for (pid, tid, phone, fn, ln, consent, created_at) in preview]

    label = "все записи"
    if args.strip().lower() == "today":
        label = "сегодня (UTC)"
    elif args.strip():
        label = f"диапазон: {args.strip()} (UTC)"

    text = f"Фильтр: <b>{label}</b>\nВсего участников: <b>{cnt}</b>\n"
    text += "Первые записи:\n" + ("\n".join(lines) if lines else "Пока пусто.")
    if cnt > len(lines):
        text += f"\n…и ещё {cnt - len(lines)}"

    await message.answer(text, parse_mode="HTML")


async def send_export(message: Message, args: str):
    from_iso, to_iso, err = range_from_args(args)
    if err:
        await message.answer(err)
        return

    rows = await fetch_participants(from_iso, to_iso)
    if not rows:
        await message.answer("По этому фильтру нет участников.")
        return

    suffix = "all"
    if args.strip().lower() == "today":
        suffix = "today_utc"
    elif args.strip():
        suffix = args.strip().replace(" ", "_")

    await export_to_excel_and_send(message, rows, suffix)


# ---------- Public flow ----------
async def show_user_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "Нажмите кнопку «🚀 Старт», чтобы начать регистрацию.",
        reply_markup=user_start_kb()
    )


async def start(message: Message, state: FSMContext):
    # /start shows the "🚀 Старт" button for participants
    existing = await get_by_telegram_id(message.from_user.id)
    if existing:
        pid, tid, phone, fn, ln, consent, created_at = existing
        await state.clear()
        await message.answer(
            f"Вы уже зарегистрированы ✅\n"
            f"Номер участника: <b>{pid}</b>\n"
            f"Имя: {fn}\nФамилия: {ln}\nТелефон: {phone}\n\n"
            f"Нажмите «🚀 Старт», чтобы открыть начало.",
            parse_mode="HTML",
            reply_markup=user_start_kb()
        )
        return

    await show_user_start(message, state)


async def on_user_start_button(message: Message, state: FSMContext):
    # Begin the registration flow from the "🚀 Старт" button
    existing = await get_by_telegram_id(message.from_user.id)
    if existing:
        pid, tid, phone, fn, ln, consent, created_at = existing
        await state.clear()
        await message.answer(
            f"Вы уже зарегистрированы ✅\nНомер участника: <b>{pid}</b>\n\nКоманда /my — показать номер.",
            parse_mode="HTML",
            reply_markup=user_start_kb()
        )
        return

    await message.answer(
        "Перед регистрацией нужно согласие на обработку данных (телефон, имя, фамилия) "
        "для целей регистрации участника.\n\n"
        "Вы согласны?",
        reply_markup=consent_kb()
    )
    await state.set_state(Reg.waiting_consent)


async def cmd_my(message: Message, state: FSMContext):
    existing = await get_by_telegram_id(message.from_user.id)
    if not existing:
        await message.answer("Вы ещё не зарегистрированы. Нажмите /start", reply_markup=user_start_kb())
        return
    pid, tid, phone, fn, ln, consent, created_at = existing
    await message.answer(
        f"Ваш номер участника: <b>{pid}</b>\n"
        f"Имя: {fn}\nФамилия: {ln}\nТелефон: {phone}",
        parse_mode="HTML",
        reply_markup=user_start_kb()
    )
    await state.clear()


async def cmd_reset(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Сбросил текущий шаг. Нажмите «🚀 Старт» или /start.", reply_markup=user_start_kb())


async def on_consent(message: Message, state: FSMContext):
    text = (message.text or "").strip()

    if text == "❌ Не согласен":
        await message.answer(
            "Понял. Без согласия я не могу зарегистрировать вас.\n"
            "Если передумаете — нажмите «🚀 Старт» или /start.",
            reply_markup=user_start_kb()
        )
        await state.clear()
        return

    if text != "✅ Согласен":
        await message.answer("Пожалуйста, выберите кнопку: ✅ Согласен или ❌ Не согласен.", reply_markup=consent_kb())
        return

    await state.update_data(consent=1)
    await message.answer(
        "Отлично. Теперь нажмите кнопку ниже, чтобы отправить номер телефона.",
        reply_markup=contact_kb()
    )
    await state.set_state(Reg.waiting_phone)


async def on_phone(message: Message, state: FSMContext):
    if not message.contact or not message.contact.phone_number:
        await message.answer(
            "Пожалуйста, отправьте номер через кнопку «📱 Отправить номер телефона».",
            reply_markup=contact_kb()
        )
        return

    if message.contact.user_id and message.contact.user_id != message.from_user.id:
        await message.answer("Пожалуйста, отправьте ваш собственный номер (через кнопку).", reply_markup=contact_kb())
        return

    phone = normalize_phone(message.contact.phone_number)

    existing = await get_by_telegram_id(message.from_user.id)
    if existing:
        pid, *_ = existing
        await message.answer(
            f"Вы уже зарегистрированы ✅\nНомер участника: <b>{pid}</b>",
            parse_mode="HTML",
            reply_markup=user_start_kb()
        )
        await state.clear()
        return

    same_phone = await get_by_phone(phone)
    if same_phone and same_phone[1] != message.from_user.id:
        pid, tid, _, fn, ln, *_ = same_phone
        await message.answer(
            "Этот номер телефона уже зарегистрирован другим участником.\n"
            "Если это ошибка — обратитесь к организатору.\n"
            f"(Номер участника по этому телефону: {pid}, имя: {fn} {ln})",
            reply_markup=user_start_kb()
        )
        await state.clear()
        return

    await state.update_data(phone=phone)
    await message.answer("Введите ваше имя:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(Reg.waiting_first_name)


async def on_first_name(message: Message, state: FSMContext):
    first_name = (message.text or "").strip()
    if not first_name or len(first_name) < 2 or len(first_name) > 50:
        await message.answer("Введите корректное имя (2–50 символов):")
        return

    await state.update_data(first_name=first_name)
    await message.answer("Введите вашу фамилию:")
    await state.set_state(Reg.waiting_last_name)


async def on_last_name(message: Message, state: FSMContext):
    last_name = (message.text or "").strip()
    if not last_name or len(last_name) < 2 or len(last_name) > 50:
        await message.answer("Введите корректную фамилию (2–50 символов):")
        return

    data = await state.get_data()
    phone = data["phone"]
    first_name = data["first_name"]
    consent = int(data.get("consent", 0))

    existing = await get_by_telegram_id(message.from_user.id)
    if existing:
        pid, *_ = existing
        await message.answer(
            f"Вы уже зарегистрированы ✅ Номер участника: <b>{pid}</b>",
            parse_mode="HTML",
            reply_markup=user_start_kb()
        )
        await state.clear()
        return

    same_phone = await get_by_phone(phone)
    if same_phone and same_phone[1] != message.from_user.id:
        await message.answer("Этот номер уже зарегистрирован. Обратитесь к организатору.", reply_markup=user_start_kb())
        await state.clear()
        return

    try:
        participant_id = await insert_participant(
            telegram_id=message.from_user.id,
            phone=phone,
            first_name=first_name,
            last_name=last_name,
            consent=consent
        )
    except aiosqlite.IntegrityError:
        existing = await get_by_telegram_id(message.from_user.id)
        if existing:
            participant_id = existing[0]
            await message.answer(
                f"Вы уже зарегистрированы ✅ Номер участника: <b>{participant_id}</b>",
                parse_mode="HTML",
                reply_markup=user_start_kb()
            )
        else:
            await message.answer("Не удалось зарегистрировать (возможно, номер уже занят). Обратитесь к организатору.",
                                 reply_markup=user_start_kb())
        await state.clear()
        return

    await message.answer(
        "Готово! Вы зарегистрированы ✅\n"
        f"Ваш номер участника: <b>{participant_id}</b>\n\n"
        "Команда /my — показать мой номер.",
        parse_mode="HTML",
        reply_markup=user_start_kb()
    )
    await state.clear()


# ---------- Admin handlers ----------
async def cmd_admin(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    await state.clear()
    await message.answer("Админ-меню:", reply_markup=admin_kb())


async def admin_close_menu(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer("Меню закрыто.", reply_markup=ReplyKeyboardRemove())


# Admin commands
async def cmd_list(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    args = message.text.replace("/list", "", 1).strip()
    await send_list(message, args)


async def cmd_export(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Команда доступна только администратору.")
        return
    args = message.text.replace("/export", "", 1).strip()
    await send_export(message, args)


# Admin menu buttons
async def admin_menu_list(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await message.answer("Выберите фильтр для списка:", reply_markup=admin_filter_kb())
    await state.set_state(AdminFSM.list_wait_from)


async def admin_menu_export(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await message.answer("Выберите фильтр для экспорта:", reply_markup=admin_filter_kb())
    await state.set_state(AdminFSM.export_wait_from)


async def admin_menu_export_today(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()
    await send_export(message, "today")
    await message.answer("Админ-меню:", reply_markup=admin_kb())


async def admin_list_filter_step(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()

    if t == "⬅️ Назад":
        await state.clear()
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Все":
        await state.clear()
        await send_list(message, "")
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Сегодня":
        await state.clear()
        await send_list(message, "today")
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Диапазон дат":
        await message.answer("Введите дату ОТ в формате YYYY-MM-DD (UTC):", reply_markup=admin_back_kb())
        await state.set_state(AdminFSM.list_wait_to)
        await state.update_data(list_step="from")
        return

    from_iso, to_iso, err = range_from_args(t)
    if err:
        await message.answer("Не понял. Нажми кнопку фильтра или введи: YYYY-MM-DD YYYY-MM-DD", reply_markup=admin_filter_kb())
        return

    await state.clear()
    await send_list(message, t)
    await message.answer("Админ-меню:", reply_markup=admin_kb())


async def admin_list_range_collect(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()
    if t == "⬅️ Назад":
        await state.clear()
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    data = await state.get_data()
    step = data.get("list_step")

    if step == "from":
        d1 = parse_ymd(t)
        if not d1:
            await message.answer("Неверная дата. Формат YYYY-MM-DD. Введите дату ОТ ещё раз:")
            return
        await state.update_data(list_from=d1.isoformat(), list_step="to")
        await message.answer("Введите дату ДО в формате YYYY-MM-DD (UTC):")
        return

    d2 = parse_ymd(t)
    if not d2:
        await message.answer("Неверная дата. Формат YYYY-MM-DD. Введите дату ДО ещё раз:")
        return

    d1 = parse_ymd(data["list_from"])
    args = f"{d1.isoformat()} {d2.isoformat()}"
    await state.clear()
    await send_list(message, args)
    await message.answer("Админ-меню:", reply_markup=admin_kb())


async def admin_export_filter_step(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()

    if t == "⬅️ Назад":
        await state.clear()
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Все":
        await state.clear()
        await send_export(message, "")
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Сегодня":
        await state.clear()
        await send_export(message, "today")
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t == "Диапазон дат":
        await message.answer("Введите дату ОТ в формате YYYY-MM-DD (UTC):", reply_markup=admin_back_kb())
        await state.set_state(AdminFSM.export_wait_to)
        await state.update_data(export_step="from")
        return

    from_iso, to_iso, err = range_from_args(t)
    if err:
        await message.answer("Не понял. Нажми кнопку фильтра или введи: YYYY-MM-DD YYYY-MM-DD", reply_markup=admin_filter_kb())
        return

    await state.clear()
    await send_export(message, t)
    await message.answer("Админ-меню:", reply_markup=admin_kb())


async def admin_export_range_collect(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()
    if t == "⬅️ Назад":
        await state.clear()
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    data = await state.get_data()
    step = data.get("export_step")

    if step == "from":
        d1 = parse_ymd(t)
        if not d1:
            await message.answer("Неверная дата. Формат YYYY-MM-DD. Введите дату ОТ ещё раз:")
            return
        await state.update_data(export_from=d1.isoformat(), export_step="to")
        await message.answer("Введите дату ДО в формате YYYY-MM-DD (UTC):")
        return

    d2 = parse_ymd(t)
    if not d2:
        await message.answer("Неверная дата. Формат YYYY-MM-DD. Введите дату ДО ещё раз:")
        return

    d1 = parse_ymd(data["export_from"])
    args = f"{d1.isoformat()} {d2.isoformat()}"
    await state.clear()
    await send_export(message, args)
    await message.answer("Админ-меню:", reply_markup=admin_kb())


# --- Reset with password ---
async def admin_menu_reset(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if not RESET_PASSWORD:
        await message.answer(
            "RESET_PASSWORD не задан в Railway Variables.\n"
            "Добавь переменную RESET_PASSWORD и попробуй снова.",
            reply_markup=admin_kb()
        )
        await state.clear()
        return

    await message.answer("Введите пароль для ресета базы:", reply_markup=admin_back_kb())
    await state.set_state(AdminFSM.reset_wait_password)


async def admin_reset_password_step(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()

    if t == "⬅️ Назад":
        await state.clear()
        await message.answer("Админ-меню:", reply_markup=admin_kb())
        return

    if t != RESET_PASSWORD:
        await message.answer("Неверный пароль. Попробуйте ещё раз или нажмите «⬅️ Назад».")
        return

    await message.answer(
        "⚠️ Пароль верный.\n"
        "Это удалит ВСЕХ участников и сбросит нумерацию обратно с 1.\n\n"
        "Подтвердить?",
        reply_markup=admin_reset_confirm_kb()
    )
    await state.set_state(AdminFSM.reset_confirm)


async def admin_reset_confirm(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    t = (message.text or "").strip()

    if t == "❌ Отмена":
        await state.clear()
        await message.answer("Ок, отменил.", reply_markup=admin_kb())
        return

    if t != "✅ Да, стереть всё":
        await message.answer("Выберите кнопку: ✅ Да, стереть всё или ❌ Отмена", reply_markup=admin_reset_confirm_kb())
        return

    await reset_database()
    await state.clear()
    await message.answer("Готово ✅ База очищена, нумерация начнётся заново с 1.", reply_markup=admin_kb())


# ---------- Main ----------
async def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN is not set. Add it in Railway -> Variables.")

    await init_db()

    bot = Bot(BOT_TOKEN)
    dp = Dispatcher()

    # Public
    dp.message.register(start, CommandStart())
    dp.message.register(on_user_start_button, lambda m: (m.text or "") == "🚀 Старт")
    dp.message.register(cmd_my, Command("my"))
    dp.message.register(cmd_reset, Command("reset"))

    # Admin commands
    dp.message.register(cmd_admin, Command("admin"))
    dp.message.register(cmd_list, Command("list"))
    dp.message.register(cmd_export, Command("export"))

    # Admin menu buttons (texts)
    dp.message.register(admin_close_menu, lambda m: is_admin(m.from_user.id) and (m.text or "") == "⬅️ Закрыть меню")
    dp.message.register(admin_menu_list, lambda m: is_admin(m.from_user.id) and (m.text or "") == "📋 Список")
    dp.message.register(admin_menu_export, lambda m: is_admin(m.from_user.id) and (m.text or "") == "📤 Экспорт")
    dp.message.register(admin_menu_export_today, lambda m: is_admin(m.from_user.id) and (m.text or "") == "📤 Экспорт сегодня")
    dp.message.register(admin_menu_reset, lambda m: is_admin(m.from_user.id) and (m.text or "") == "🧹 Ресет базы")

    # Admin FSM
    dp.message.register(admin_reset_password_step, AdminFSM.reset_wait_password)
    dp.message.register(admin_reset_confirm, AdminFSM.reset_confirm)

    dp.message.register(admin_list_filter_step, AdminFSM.list_wait_from)
    dp.message.register(admin_list_range_collect, AdminFSM.list_wait_to)

    dp.message.register(admin_export_filter_step, AdminFSM.export_wait_from)
    dp.message.register(admin_export_range_collect, AdminFSM.export_wait_to)

    # Registration FSM
    dp.message.register(on_consent, Reg.waiting_consent)
    dp.message.register(on_phone, Reg.waiting_phone)
    dp.message.register(on_first_name, Reg.waiting_first_name)
    dp.message.register(on_last_name, Reg.waiting_last_name)

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
